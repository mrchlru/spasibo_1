from fastapi import APIRouter, Depends, HTTPException, Response, Query, status
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.exc import IntegrityError
from sqlalchemy import select, func, union_all, literal, case
from sqlalchemy.orm import selectinload
from typing import List, Optional
from datetime import date, datetime, timedelta
from zoneinfo import ZoneInfo
import crud
import schemas
import models
import io
from fastapi.responses import StreamingResponse
from dependencies import get_current_admin_user
from database import get_db
from config import settings
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

router = APIRouter(
    prefix="/admin",
    tags=["admin"],
    dependencies=[Depends(get_current_admin_user)]
)

class AddPointsRequest(BaseModel):
    amount: int
class AddTicketsRequest(BaseModel):
    amount: int

@router.post("/add-points")
async def add_points(request: AddPointsRequest, db: AsyncSession = Depends(get_db)):
    await crud.add_points_to_all_users(db, amount=request.amount)
    return {"detail": f"Successfully added {request.amount} points to all users"}

@router.post("/add-tickets")
async def add_tickets(request: AddTicketsRequest, db: AsyncSession = Depends(get_db)):
    await crud.add_tickets_to_all_users(db, amount=request.amount)
    return {"detail": f"Successfully added {request.amount} tickets to all users"}

@router.post("/market-items", response_model=schemas.MarketItemResponse)
async def create_new_market_item(
    item: schemas.MarketItemCreate, 
    db: AsyncSession = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    try:
        new_item = await crud.admin_create_market_item(db=db, item=item)
        return new_item
    except IntegrityError:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Один или несколько из предоставленных кодов уже существуют. Убедитесь, что все коды уникальны."
        )

@router.put("/market-items/{item_id}", response_model=schemas.MarketItemResponse)
async def update_market_item_route(
    item_id: int,
    item_data: schemas.MarketItemUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    print(f"--- [ROUTER UPDATE {item_id}] Получен PUT запрос ---")
    print(f"--- [ROUTER UPDATE {item_id}] Данные от фронтенда: {item_data.model_dump(exclude_unset=True)}")
    try:
        updated_item = await crud.admin_update_market_item(db, item_id, item_data)
        if not updated_item:
             print(f"--- [ROUTER UPDATE {item_id}] CRUD вернул None (Товар не найден) ---")
             raise HTTPException(status_code=404, detail="Товар не найден")
        print(f"--- [ROUTER UPDATE {item_id}] CRUD успешно вернул обновленный товар: {updated_item.name} ---")
        return updated_item
    except IntegrityError:
        print(f"--- [ROUTER UPDATE {item_id}] Произошла ошибка IntegrityError ---")
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Один или несколько из предоставленных кодов уже существуют."
        )
    except ValueError as e:
        print(f"--- [ROUTER UPDATE {item_id}] Произошла ошибка ValueError: {e} ---")
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        print(f"--- [ROUTER UPDATE {item_id}] НЕОЖИДАННАЯ ОШИБКА: {type(e).__name__} - {e} ---")
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Произошла внутренняя ошибка сервера при обновлении."
        )

@router.delete("/market-items/{item_id}", status_code=204)
async def archive_item_route(item_id: int, db: AsyncSession = Depends(get_db)):
    success = await crud.archive_market_item(db, item_id)
    if not success:
        raise HTTPException(status_code=404, detail="Item not found")
    return Response(status_code=204)
    
@router.get("/market-items/archived", response_model=list[schemas.MarketItemResponse])
async def get_archived_items_route(db: AsyncSession = Depends(get_db)):
    items = await crud.get_archived_items(db)
    return items

@router.get("/market-items", response_model=List[schemas.MarketItemResponse])
async def get_all_active_items_route(db: AsyncSession = Depends(get_db)):
    return await crud.get_active_items(db)

@router.post("/market-items/{item_id}/restore", response_model=schemas.MarketItemResponse)
async def restore_item_route(
    item_id: int, 
    db: AsyncSession = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    restored_item = await crud.admin_restore_market_item(db, item_id)
    if not restored_item:
        raise HTTPException(status_code=404, detail="Товар не найден")
    return restored_item

@router.post("/reset-balances")
async def reset_balances_route(db: AsyncSession = Depends(get_db)):
    await crud.reset_balances(db)
    return {"detail": "Balances reset successfully"}

@router.post("/reset-daily-transfer-limits")
async def reset_daily_transfer_limits_route(db: AsyncSession = Depends(get_db)):
    await crud.reset_daily_transfer_limits(db)
    return {"detail": "Лимиты на отправку спасибок успешно сброшены у всех пользователей"}

@router.get("/users", response_model=List[schemas.UserResponse])
async def get_all_users_for_admin_route(db: AsyncSession = Depends(get_db)):
    return await crud.get_all_users_for_admin(db)

@router.get("/users/pending", response_model=List[schemas.UserResponse])
async def get_pending_users_route(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(models.User).where(models.User.status == 'pending').order_by(models.User.registration_date.desc())
    )
    return result.scalars().all()

@router.post("/users/{user_id}/approve", response_model=schemas.ApproveUserRegistrationResponse)
async def approve_user_registration_route(
    user_id: int,
    admin_user: models.User = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    updated_user = await crud.update_user_status(db, user_id, "approved")
    if not updated_user:
        raise HTTPException(status_code=404, detail="User not found")

    gen_login = getattr(updated_user, "_generated_login", None) or updated_user.login
    gen_password = getattr(updated_user, "_generated_password", None) or updated_user.password_plain
    created_now = getattr(updated_user, "_credentials_created_this_call", False)

    return schemas.ApproveUserRegistrationResponse(
        user=schemas.UserResponse.model_validate(updated_user),
        login=gen_login,
        password=gen_password,
        credentials_generated=created_now,
        credentials_active=True,
    )

@router.post("/users/{user_id}/reject", response_model=schemas.UserResponse)
async def reject_user_registration_route(
    user_id: int,
    admin_user: models.User = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    updated_user = await crud.update_user_status(db, user_id, "rejected")
    if not updated_user:
        raise HTTPException(status_code=404, detail="User not found")
    return updated_user

@router.get("/local-purchases/pending")
async def get_pending_local_purchases_route(db: AsyncSession = Depends(get_db)):
    """Получает список локальных покупок, ожидающих согласования"""
    from sqlalchemy import select
    from models import LocalGift
    result = await db.execute(
        select(LocalGift).where(LocalGift.status == 'pending').order_by(LocalGift.created_at.desc())
    )
    local_purchases = result.scalars().all()
    return [
        {
            "id": lp.id,
            "user_id": lp.user_id,
            "user_name": f"{lp.user.first_name} {lp.user.last_name}",
            "item_id": lp.item_id,
            "item_name": lp.item.name,
            "city": lp.city,
            "website_url": lp.website_url,
            "reserved_amount": lp.reserved_amount,
            "created_at": lp.created_at.isoformat() if lp.created_at else None,
            "status": lp.status
        }
        for lp in local_purchases
    ]

@router.get("/profile-updates/pending")
async def get_pending_profile_updates_route(db: AsyncSession = Depends(get_db)):
    """Получает список запросов на изменение профиля, ожидающих согласования"""
    from sqlalchemy import select
    from models import PendingUpdate
    result = await db.execute(
        select(PendingUpdate)
        .options(selectinload(PendingUpdate.user))
        .where(PendingUpdate.status == 'pending')
        .order_by(PendingUpdate.created_at.desc())
    )
    pending_updates = result.scalars().all()
    return [
        {
            "id": pu.id,
            "user_id": pu.user_id,
            "user_name": f"{pu.user.first_name} {pu.user.last_name}",
            "old_data": pu.old_data,
            "new_data": pu.new_data,
            "created_at": pu.created_at.isoformat() if pu.created_at else None,
            "status": pu.status
        }
        for pu in pending_updates
    ]

@router.post("/local-purchases/{purchase_id}/approve")
async def approve_local_purchase_route(
    purchase_id: int,
    db: AsyncSession = Depends(get_db),
    admin_user: models.User = Depends(get_current_admin_user)
):
    """Одобряет локальную покупку"""
    result = await crud.process_local_gift_approval(db, purchase_id, 'approve')
    if result is None:
        raise HTTPException(status_code=404, detail="Покупка не найдена или уже обработана")
    return result

@router.post("/local-purchases/{purchase_id}/reject")
async def reject_local_purchase_route(
    purchase_id: int,
    db: AsyncSession = Depends(get_db),
    admin_user: models.User = Depends(get_current_admin_user)
):
    """Отклоняет локальную покупку"""
    result = await crud.process_local_gift_approval(db, purchase_id, 'reject')
    if result is None:
        raise HTTPException(status_code=404, detail="Покупка не найдена или уже обработана")
    return result

@router.post("/profile-updates/{update_id}/approve")
async def approve_profile_update_route(
    update_id: int,
    db: AsyncSession = Depends(get_db),
    admin_user: models.User = Depends(get_current_admin_user)
):
    """Одобряет запрос на изменение профиля"""
    user, status = await crud.process_profile_update(db, update_id, 'approve')
    if user is None:
        raise HTTPException(status_code=404, detail="Запрос не найден или уже обработан")
    return {"status": status, "user": schemas.UserResponse.model_validate(user)}

@router.post("/profile-updates/{update_id}/reject")
async def reject_profile_update_route(
    update_id: int,
    db: AsyncSession = Depends(get_db),
    admin_user: models.User = Depends(get_current_admin_user)
):
    """Отклоняет запрос на изменение профиля"""
    user, status = await crud.process_profile_update(db, update_id, 'reject')
    if user is None:
        raise HTTPException(status_code=404, detail="Запрос не найден или уже обработан")
    return {"status": status, "user": schemas.UserResponse.model_validate(user)}

@router.put("/users/{user_id}", response_model=schemas.UserResponse)
async def admin_update_user_route(user_id: int, user_data: schemas.AdminUserUpdate, admin_user: models.User = Depends(get_current_admin_user), db: AsyncSession = Depends(get_db)):
    try:
        updated_user = await crud.admin_update_user(db, user_id, user_data, admin_user)
        if not updated_user:
            raise HTTPException(status_code=404, detail="User not found")
        return updated_user
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.delete("/users/{user_id}", status_code=204)
async def admin_delete_user_route(user_id: int, admin_user: models.User = Depends(get_current_admin_user), db: AsyncSession = Depends(get_db)):
    success = await crud.admin_delete_user(db, user_id, admin_user)
    if not success:
        raise HTTPException(status_code=404, detail="User not found")
    return Response(status_code=204)

@router.post("/users/{user_id}/set-credentials", response_model=schemas.SetUserCredentialsResponse)
async def set_user_credentials_route(
    user_id: int,
    credentials: schemas.SetUserCredentialsRequest,
    admin_user: models.User = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    try:
        updated_user = await crud.set_user_credentials(
            db, 
            user_id, 
            credentials.login, 
            credentials.password
        )
        
        return schemas.SetUserCredentialsResponse(
            message="Учетные данные успешно установлены",
            login=updated_user.login,
            user_id=updated_user.id
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        print(f"Ошибка при установке учетных данных: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось установить учетные данные"
        )

class ChangeUserPasswordRequest(BaseModel):
    new_password: str

@router.post("/users/{user_id}/change-password", response_model=schemas.UserResponse)
async def admin_change_user_password_route(
    user_id: int,
    password_data: ChangeUserPasswordRequest,
    admin_user: models.User = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    """Изменяет пароль пользователя от имени администратора."""
    try:
        updated_user = await crud.admin_change_user_password(
            db,
            user_id,
            password_data.new_password,
            admin_user
        )
        return updated_user
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        print(f"Ошибка при изменении пароля: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось изменить пароль"
        )

@router.delete("/users/{user_id}/password", response_model=schemas.UserResponse)
async def admin_delete_user_password_route(
    user_id: int,
    admin_user: models.User = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    """Удаляет пароль пользователя, отключая вход через браузер."""
    try:
        updated_user = await crud.admin_delete_user_password(
            db,
            user_id,
            admin_user
        )
        return updated_user
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        print(f"Ошибка при удалении пароля: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось удалить пароль"
        )

@router.post("/users/bulk-send-credentials", response_model=schemas.BulkSendCredentialsResponse)
async def bulk_send_credentials_route(
    request: schemas.BulkSendCredentialsRequest,
    admin_user: models.User = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    try:
        result = await crud.bulk_send_credentials(
            db,
            custom_message=request.message or "",
            include_active=request.include_active,
            include_blocked=request.include_blocked,
            regenerate_existing=request.regenerate_existing
        )
        
        return schemas.BulkSendCredentialsResponse(
            message=f"Рассылка завершена. Сгенерировано учетных данных: {result['credentials_generated']}, отправлено сообщений: {result['messages_sent']}",
            total_users=result['total_users'],
            credentials_generated=result['credentials_generated'],
            messages_sent=result['messages_sent'],
            failed_users=result['failed_users']
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        print(f"Ошибка при массовой рассылке учетных данных: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось выполнить массовую рассылку"
        )


@router.get("/users/broadcast-email/preview", response_model=schemas.BroadcastEmailPreviewResponse)
async def broadcast_email_preview_route(
    only_browser_users: bool = Query(
        True,
        description="Считать только пользователей с доступом через браузер (не заблокированы, одобрены)",
    ),
    db: AsyncSession = Depends(get_db),
):
    """Оценка числа получателей по email и в Telegram (независимо от выбранных каналов при отправке)."""
    n_email = await crud.count_broadcast_email_recipients(db, only_browser_users)
    n_tg = await crud.count_broadcast_telegram_recipients(db, only_browser_users)
    return schemas.BroadcastEmailPreviewResponse(
        recipient_count_email=n_email,
        recipient_count_telegram=n_tg,
    )


@router.get("/users/broadcast/eligible", response_model=schemas.BroadcastEligibleResponse)
async def broadcast_eligible_users_route(
    only_browser_users: bool = Query(
        True,
        description="Только пользователи с включённым входом через браузер",
    ),
    db: AsyncSession = Depends(get_db),
):
    """Список пользователей-кандидатов на рассылку (для UI выбора получателей)."""
    users = await crud.fetch_broadcast_eligible_users(db, only_browser_users)
    available_email = sum(1 for u in users if u.get("email_available"))
    available_telegram = sum(1 for u in users if u.get("telegram_available"))
    no_dialog = sum(1 for u in users if u.get("telegram_no_dialog"))
    return schemas.BroadcastEligibleResponse(
        only_browser_users=only_browser_users,
        users=[schemas.BroadcastEligibleUser(**u) for u in users],
        total=len(users),
        available_email=available_email,
        available_telegram=available_telegram,
        telegram_no_dialog_count=no_dialog,
    )


@router.post("/users/broadcast/export-report")
async def broadcast_export_report_route(
    payload: schemas.BroadcastReportExportRequest,
):
    """Возвращает Excel-файл с отчётом о последней рассылке.

    Сервер не хранит результаты рассылки между запросами, поэтому фронт
    отправляет уже полученный отчёт целиком. На лист попадают: ФИО, отдел,
    должность, телефон, email, telegram_id, канал, статус и человекочитаемая
    причина (если сообщение не доставлено).
    """
    moscow_tz = ZoneInfo("Europe/Moscow")
    now_moscow = datetime.now(moscow_tz)

    rows: list[dict[str, object]] = []
    for r in payload.recipients:
        if r.ok:
            status_text = "Доставлено"
        elif r.skipped:
            status_text = "Пропущено"
        else:
            status_text = "Не доставлено"
        contact = ""
        if r.channel == "email":
            contact = r.target or r.email or ""
        else:
            contact = (
                f"tg:{r.telegram_id}" if r.telegram_id is not None else (r.target or "")
            )
        full_name = (
            f"{r.last_name} {r.first_name}".strip() or r.name or f"id:{r.user_id}"
        )
        rows.append(
            {
                "ID": r.user_id,
                "Фамилия": r.last_name or "",
                "Имя": r.first_name or "",
                "ФИО": full_name,
                "Отдел": r.department or "",
                "Должность": r.position or "",
                "Телефон": r.phone or "",
                "Email": r.email or "",
                "Telegram ID": r.telegram_id if r.telegram_id is not None else "",
                "Канал": "Email" if r.channel == "email" else "Telegram",
                "Контакт": contact,
                "Статус": status_text,
                "Причина": r.reason or ("" if r.ok else "—"),
                "Код ошибки": r.error_code or "",
                "Подробности": (r.detail or "")[:500],
            }
        )

    output = io.BytesIO()
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Отчёт по рассылке"

    # Метаданные о рассылке — сводка наверху отдельным маленьким блоком
    ws.append(["Тема", payload.subject or ""])
    ws.append(["Дата формирования (МСК)", now_moscow.strftime("%Y-%m-%d %H:%M")])
    ws.append(
        ["Email доставлено", f"{payload.sent_ok_email}/{payload.recipient_count_email}"]
    )
    ws.append(
        [
            "Telegram доставлено",
            f"{payload.sent_ok_telegram}/{payload.recipient_count_telegram}",
        ]
    )
    ws.append([])

    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
    else:
        ws.append(["Нет данных для экспорта"])

    # Авто-ширина колонок (грубо, но достаточно для читаемости)
    for column_cells in ws.columns:
        max_length = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    workbook.save(output)
    output.seek(0)

    filename = f"broadcast_report_{now_moscow.strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/users/broadcast-email", response_model=schemas.BroadcastEmailResponse)
async def broadcast_email_route(
    request: schemas.BroadcastEmailRequest,
    db: AsyncSession = Depends(get_db),
):
    """Рассылка в email, Telegram или оба канала: одобренные пользователи, не заблокированы.

    Если в теле передан ``user_ids`` — отправка ограничена этим списком.
    """
    if request.send_email:
        smtp_user = getattr(settings, "SMTP_USERNAME", None) or ""
        smtp_pass = getattr(settings, "SMTP_PASSWORD", None) or ""
        if not str(smtp_user).strip() or not str(smtp_pass).strip():
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="SMTP не настроен (SMTP_USERNAME / SMTP_PASSWORD). Проверьте переменные окружения.",
            )
    if request.send_telegram:
        token = getattr(settings, "TELEGRAM_BOT_TOKEN", None) or ""
        if not str(token).strip():
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="Telegram бот не настроен (TELEGRAM_BOT_TOKEN).",
            )
    result = await crud.broadcast_announcement_to_users(
        db,
        subject=request.subject,
        body_plain=request.body,
        only_browser_users=request.only_browser_users,
        append_login_url=request.append_login_url,
        do_send_email=request.send_email,
        do_send_telegram=request.send_telegram,
        user_ids=request.user_ids,
    )
    parts: list[str] = []
    if request.send_email:
        parts.append(
            f"Email: {result['sent_ok_email']}/{result['recipient_count_email']}",
        )
    if request.send_telegram:
        parts.append(
            f"Telegram: {result['sent_ok_telegram']}/{result['recipient_count_telegram']}",
        )
    msg = "; ".join(parts) if parts else "Ничего не отправлено"
    return schemas.BroadcastEmailResponse(
        message=msg,
        recipient_count_email=result["recipient_count_email"],
        recipient_count_telegram=result["recipient_count_telegram"],
        sent_ok_email=result["sent_ok_email"],
        sent_ok_telegram=result["sent_ok_telegram"],
        failed=[schemas.BroadcastEmailFailedItem(**f) for f in result["failed"]],
        recipients=[
            schemas.BroadcastRecipientReport(**r) for r in result.get("recipients", [])
        ],
    )


@router.get("/statistics/general", response_model=schemas.GeneralStatsResponse)
async def get_general_statistics_route(start_date: Optional[date] = Query(None), end_date: Optional[date] = Query(None), db: AsyncSession = Depends(get_db)):
    stats = await crud.get_general_statistics(db=db, start_date=start_date, end_date=end_date)
    return stats

@router.get("/statistics/hourly_activity", response_model=schemas.HourlyActivityStats)
async def get_hourly_activity(start_date: Optional[date] = Query(None), end_date: Optional[date] = Query(None), db: AsyncSession = Depends(get_db)):
    stats = await crud.get_hourly_activity_stats(db, start_date=start_date, end_date=end_date)
    return {"hourly_stats": stats}

@router.get("/statistics/login_activity", response_model=schemas.LoginActivityStats)
async def get_login_activity(start_date: Optional[date] = Query(None), end_date: Optional[date] = Query(None), db: AsyncSession = Depends(get_db)):
    stats = await crud.get_login_activity_stats(db, start_date=start_date, end_date=end_date)
    return {"hourly_stats": stats}

@router.get("/statistics/active_user_ratio", response_model=schemas.ActiveUserRatioStats)
async def get_active_user_ratio_route(db: AsyncSession = Depends(get_db)):
    ratio_data = await crud.get_active_user_ratio(db)
    return ratio_data

@router.get("/statistics/user_engagement", response_model=schemas.UserEngagementStats)
async def get_user_engagement(db: AsyncSession = Depends(get_db)):
    engagement_data = await crud.get_user_engagement_stats(db)
    top_senders_schema = [
        {"user": row[0], "count": row.sent_count} 
        for row in engagement_data["top_senders"]
    ]
    top_receivers_schema = [
        {"user": row[0], "count": row.received_count} 
        for row in engagement_data["top_receivers"]
    ]
    return {"top_senders": top_senders_schema, "top_receivers": top_receivers_schema}

@router.get("/statistics/popular_items", response_model=schemas.PopularItemsStats)
async def get_popular_items(db: AsyncSession = Depends(get_db)):
    items_data = await crud.get_popular_items_stats(db)
    popular_items_schema = [
        {"item": row[0], "purchase_count": row.purchase_count} 
        for row in items_data
    ]
    return {"items": popular_items_schema}

@router.get("/statistics/inactive_users", response_model=schemas.InactiveUsersStats)
async def get_inactive_users_list(db: AsyncSession = Depends(get_db)):
    inactive_users = await crud.get_inactive_users(db)
    return {"users": inactive_users}

@router.get("/statistics/total_balance", response_model=schemas.TotalBalanceStats)
async def get_economy_total_balance(db: AsyncSession = Depends(get_db)):
    total_balance = await crud.get_total_balance(db)
    return {"total_balance": total_balance}

@router.get("/statistics/average_session_duration", response_model=schemas.AverageSessionDurationStats)
async def get_average_session_duration_route(
    start_date: Optional[date] = Query(None), 
    end_date: Optional[date] = Query(None), 
    db: AsyncSession = Depends(get_db)
):
    stats = await crud.get_average_session_duration(db, start_date=start_date, end_date=end_date)
    return stats

@router.get("/statistics/user_engagement/export")
async def export_user_engagement(db: AsyncSession = Depends(get_db)):
    engagement_data = await crud.get_user_engagement_stats(db)
    top_senders = engagement_data["top_senders"]
    top_receivers = engagement_data["top_receivers"]

    workbook = Workbook()
    
    ws_senders = workbook.active
    ws_senders.title = "Топ Донаторы"
    ws_senders.append(["#", "Имя", "Фамилия", "Должность", "Отдел", "Отправлено"])
    for i, sender_row in enumerate(top_senders, 1):
        user, count = sender_row[0], sender_row.sent_count
        ws_senders.append([i, user.first_name, user.last_name, user.position, user.department, count])

    ws_receivers = workbook.create_sheet("Топ Инфлюенсеры")
    ws_receivers.append(["#", "Имя", "Фамилия", "Должность", "Отдел", "Получено"])
    for i, receiver_row in enumerate(top_receivers, 1):
        user, count = receiver_row[0], receiver_row.received_count
        ws_receivers.append([i, user.first_name, user.last_name, user.position, user.department, count])

    file_stream = io.BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)

    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=leaders_report.xlsx"}
    )

@router.get("/statistics/export/consolidated")
async def export_consolidated_report(
    start_date: Optional[date] = Query(None), 
    end_date: Optional[date] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    if end_date is None: end_date = datetime.utcnow().date()
    if start_date is None: start_date = end_date - timedelta(days=30)

    general_stats = await crud.get_general_statistics(db, start_date, end_date)
    avg_session_stats = await crud.get_average_session_duration(db, start_date, end_date)
    engagement_data = await crud.get_user_engagement_stats(db, start_date, end_date)
    popular_items_data = await crud.get_popular_items_stats(db, start_date, end_date)
    inactive_users_data = await crud.get_inactive_users(db, start_date, end_date)
    
    general_stats['average_session_duration_minutes'] = avg_session_stats['average_duration_minutes']

    moscow_tz = ZoneInfo("Europe/Moscow")
    output = io.BytesIO()

    general_stats_translation = {
        "new_users_count": "Всего пользователей",
        "active_users_count": "Активные пользователи",
        "transactions_count": "Всего транзакций",
        "store_purchases_count": "Покупок в магазине",
        "total_turnover": "Оборот 'спасибок'",
        "total_store_spent": "Потрачено в магазине",
        "average_session_duration_minutes": "Среднее время сессии (мин)"
    }

    translated_stats = {
        general_stats_translation.get(key, key): value
        for key, value in general_stats.items()
    }

    workbook = Workbook()
    ws_general = workbook.active
    ws_general.title = "Общая статистика"
    ws_general.append(["Метрика", "Значение"])
    for metric, value in translated_stats.items():
        ws_general.append([metric, value])

    senders_list = [
        {"#": i, "Имя": row[0].first_name, "Фамилия": row[0].last_name, "Должность": row[0].position, "Отправлено": row.sent_count}
        for i, row in enumerate(engagement_data["top_senders"], 1)
    ]
    _append_dict_rows_sheet(workbook, "Топ Донаторы", senders_list)

    receivers_list = [
        {"#": i, "Имя": row[0].first_name, "Фамилия": row[0].last_name, "Должность": row[0].position, "Получено": row.received_count}
        for i, row in enumerate(engagement_data["top_receivers"], 1)
    ]
    _append_dict_rows_sheet(workbook, "Топ Инфлюенсеры", receivers_list)

    items_list = [
        {"#": i, "Название товара": row[0].name, "Цена": row[0].price, "Кол-во покупок": row.purchase_count}
        for i, row in enumerate(popular_items_data, 1)
    ]
    _append_dict_rows_sheet(workbook, "Популярные товары", items_list)

    inactive_list = [
        {
            "Имя": user.first_name,
            "Фамилия": user.last_name,
            "Должность": user.position,
            "Отдел": user.department,
            "Дата регистрации": _format_dt_moscow(user.registration_date, moscow_tz),
            "Последний вход": _format_dt_moscow(user.last_login_date, moscow_tz),
        }
        for user in inactive_users_data
    ]
    _append_dict_rows_sheet(workbook, "Неактивные пользователи", inactive_list)

    workbook.save(output)
    output.seek(0)
    filename = f"consolidated_report_{start_date}_to_{end_date}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
    
@router.get("/users/export")
async def export_all_users(db: AsyncSession = Depends(get_db)):
    all_users = await crud.get_all_users_for_admin(db)

    moscow_tz = ZoneInfo("Europe/Moscow")

    users_list = [
        {
            "ID": user.id,
            "Telegram ID": user.telegram_id,
            "Имя": user.first_name,
            "Фамилия": user.last_name,
            "Username": user.username,
            "Отдел": user.department,
            "Должность": user.position,
            "Баланс": user.balance,
            "Билеты": user.tickets,
            "Статус": user.status,
            "Админ": "Да" if user.is_admin else "Нет",
            "Дата регистрации": _format_dt_moscow(user.registration_date, moscow_tz),
            "Последний вход": _format_dt_moscow(user.last_login_date, moscow_tz),
        }
        for user in all_users
    ]

    output = io.BytesIO()
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Все пользователи"
    _write_dict_rows_on_sheet(ws, users_list)
    workbook.save(output)
    output.seek(0)

    filename = f"all_users_{datetime.utcnow().date()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.delete("/market-items/{item_id}/permanent", status_code=status.HTTP_204_NO_CONTENT)
async def delete_item_permanently_route(
    item_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    try:
        success = await crud.admin_delete_item_permanently(db, item_id)
        if not success:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Товар не найден")
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    
    return Response(status_code=status.HTTP_204_NO_CONTENT)

@router.get("/statix-bonus", response_model=schemas.StatixBonusItemResponse)
async def get_statix_bonus_settings(
    db: AsyncSession = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    item = await crud.get_statix_bonus_item(db)
    if not item:
        default_item = {
            "name": "Бонусы Statix",
            "description": "Покупка бонусов для платформы Statix",
            "thanks_to_statix_rate": 10,
            "min_bonus_per_step": 100,
            "max_bonus_per_step": 10000,
            "bonus_step": 100
        }
        item = await crud.create_statix_bonus_item(db, default_item)
    return item

@router.put("/statix-bonus", response_model=schemas.StatixBonusItemResponse)
async def update_statix_bonus_settings(
    item_data: schemas.StatixBonusItemUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    existing_item = await crud.get_statix_bonus_item(db)
    
    if existing_item:
        updated_item = await crud.update_statix_bonus_item(
            db, existing_item.id, item_data.model_dump(exclude_unset=True)
        )
    else:
        default_data = {
            "name": "Бонусы Statix",
            "description": "Покупка бонусов для платформы Statix",
            "thanks_to_statix_rate": 10,
            "min_bonus_per_step": 100,
            "max_bonus_per_step": 10000,
            "bonus_step": 100
        }
        update_data = {**default_data, **item_data.model_dump(exclude_unset=True)}
        updated_item = await crud.create_statix_bonus_item(db, update_data)
    
    return updated_item

@router.post("/generate-leaderboard-banners", status_code=status.HTTP_201_CREATED)
async def trigger_leaderboard_banner_generation(
    admin_user: models.User = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    try:
        await crud.generate_monthly_leaderboard_banners(db)
        return {"detail": "Баннеры рейтинга успешно сгенерированы."}
    except Exception as e:
        print(f"Ошибка при ручной генерации баннеров: {e}")
        raise HTTPException(status_code=500, detail="Не удалось сгенерировать баннеры")

@router.post("/generate-test-banners", status_code=status.HTTP_201_CREATED)
async def trigger_leaderboard_test_banner_generation(
    admin_user: models.User = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_db)
):
    try:
        await crud.generate_current_month_test_banners(db)
        return {"detail": "Тестовые баннеры (на основе ТЕКУЩЕГО месяца) успешно сгенерированы."}
    except Exception as e:
        print(f"Ошибка при ручной генерации ТЕСТОВЫХ баннеров: {e}")
        raise HTTPException(status_code=500, detail="Не удалось сгенерировать тестовые баннеры")


@router.get("/purchases/all", response_model=schemas.UnifiedPurchaseListResponse)
async def get_all_purchases(
    type: Optional[str] = Query(None, description="regular, local, statix, shared"),
    status_filter: Optional[str] = Query(None, alias="status", description="pending, approved, rejected, expired, completed"),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
):
    """Возвращает все покупки всех типов с фильтрацией и пагинацией."""
    items, total = await _get_all_purchases_from_db(db, type, status_filter, page, per_page)
    return schemas.UnifiedPurchaseListResponse(items=items, total=total)


async def _get_all_purchases_from_db(
    db: AsyncSession,
    purchase_type: Optional[str],
    status_filter: Optional[str],
    page: int,
    per_page: int,
) -> tuple[list[schemas.UnifiedPurchaseResponse], int]:
    """Собирает и объединяет все покупки из разных таблиц."""
    result_items: list[schemas.UnifiedPurchaseResponse] = []

    # --- Regular purchases ---
    if not purchase_type or purchase_type == 'regular':
        q = (
            select(models.Purchase)
            .join(models.MarketItem, models.Purchase.item_id == models.MarketItem.id)
            .join(models.User, models.Purchase.user_id == models.User.id)
            .options(selectinload(models.Purchase.user), selectinload(models.Purchase.item))
            .where(models.MarketItem.is_local_purchase == False)  # noqa: E712
            .where(models.MarketItem.is_shared_gift == False)  # noqa: E712
        )
        rows = (await db.execute(q)).scalars().all()
        for p in rows:
            is_statix = p.item and 'statix' in (p.item.name or '').lower()
            ptype = 'statix' if is_statix else 'regular'
            if purchase_type and ptype != purchase_type:
                continue
            u = p.user
            result_items.append(schemas.UnifiedPurchaseResponse(
                id=p.id,
                purchase_type=ptype,
                user_name=f"{u.first_name or ''} {u.last_name or ''}".strip(),
                user_id=p.user_id,
                item_name=p.item.name if p.item else "—",
                item_id=p.item_id,
                amount=p.item.price if p.item else 0,
                status="completed",
                created_at=p.timestamp or datetime.utcnow(),
                phone_number=u.phone_number,
                first_name=u.first_name,
                last_name=u.last_name,
                position=u.position,
                email=u.email,
                telegram_username=u.username,
            ))

    # --- Local purchases ---
    if not purchase_type or purchase_type == 'local':
        q = (
            select(models.LocalGift)
            .options(selectinload(models.LocalGift.user), selectinload(models.LocalGift.item))
        )
        rows = (await db.execute(q)).scalars().all()
        for lp in rows:
            u = lp.user
            result_items.append(schemas.UnifiedPurchaseResponse(
                id=lp.id,
                purchase_type='local',
                user_name=f"{u.first_name or ''} {u.last_name or ''}".strip(),
                user_id=lp.user_id,
                item_name=lp.item.name if lp.item else "—",
                item_id=lp.item_id,
                amount=lp.reserved_amount,
                status=lp.status,
                created_at=lp.created_at or datetime.utcnow(),
                city=lp.city,
                website_url=lp.website_url,
                phone_number=u.phone_number,
                first_name=u.first_name,
                last_name=u.last_name,
                position=u.position,
                email=u.email,
                telegram_username=u.username,
            ))

    # --- Shared gifts ---
    if not purchase_type or purchase_type == 'shared':
        q = (
            select(models.SharedGiftInvitation)
            .options(
                selectinload(models.SharedGiftInvitation.buyer),
                selectinload(models.SharedGiftInvitation.invited_user),
                selectinload(models.SharedGiftInvitation.item),
            )
        )
        rows = (await db.execute(q)).scalars().all()
        for sg in rows:
            buyer = sg.buyer
            buyer_name = f"{buyer.first_name or ''} {buyer.last_name or ''}".strip() if buyer else "—"
            invited_name = f"{sg.invited_user.first_name or ''} {sg.invited_user.last_name or ''}".strip() if sg.invited_user else ""
            result_items.append(schemas.UnifiedPurchaseResponse(
                id=sg.id,
                purchase_type='shared',
                user_name=f"{buyer_name} + {invited_name}" if invited_name else buyer_name,
                user_id=sg.buyer_id,
                item_name=sg.item.name if sg.item else "—",
                item_id=sg.item_id,
                amount=sg.item.price if sg.item else 0,
                status=sg.status,
                created_at=sg.created_at or datetime.utcnow(),
                phone_number=buyer.phone_number if buyer else None,
                first_name=buyer.first_name if buyer else None,
                last_name=buyer.last_name if buyer else None,
                position=buyer.position if buyer else None,
                email=buyer.email if buyer else None,
                telegram_username=buyer.username if buyer else None,
            ))

    # --- Status filter ---
    if status_filter:
        result_items = [i for i in result_items if i.status == status_filter]

    # --- Sort by date ---
    result_items.sort(key=lambda x: x.created_at, reverse=True)
    total = len(result_items)

    # --- Paginate ---
    offset = (page - 1) * per_page
    page_items = result_items[offset:offset + per_page]

    return page_items, total


def _format_dt_moscow(dt: Optional[datetime], tz: ZoneInfo) -> Optional[str]:
    """Форматирует aware-datetime в локальное время Москвы."""
    if dt is None:
        return None
    return dt.astimezone(tz).strftime("%Y-%m-%d %H:%M")


def _write_dict_rows_on_sheet(ws: Worksheet, rows: list[dict[str, object]]) -> None:
    """Записывает заголовки и строки из списка словарей на существующий лист."""
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])


def _append_dict_rows_sheet(workbook: Workbook, title: str, rows: list[dict[str, object]]) -> None:
    """Создаёт лист и пишет таблицу из списка словарей (ключи первой строки — заголовки)."""
    ws = workbook.create_sheet(title)
    _write_dict_rows_on_sheet(ws, rows)
