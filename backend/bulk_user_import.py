"""Парсинг Excel и массовое создание пользователей для концепций «Сердце»."""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

import models
from config import settings
from crud import (
    _ensure_unique_login,
    generate_login_from_name,
    generate_random_password,
    get_password_hash,
)
from email_service import send_credentials_to_user

logger = logging.getLogger(__name__)

TEMPLATE_HEADERS = [
    "Фамилия",
    "Имя",
    "Email",
    "Телефон",
    "Должность",
    "Ресторан",
    "Дата рождения (ДД.ММ.ГГГГ)",
]

HEADER_ALIASES: dict[str, str] = {
    "фамилия": "last_name",
    "имя": "first_name",
    "email": "email",
    "e-mail": "email",
    "почта": "email",
    "телефон": "phone_number",
    "phone": "phone_number",
    "должность": "position",
    "ресторан": "department",
    "отдел": "department",
    "department": "department",
    "дата рождения": "date_of_birth",
    "дата рождения (дд.мм.гггг)": "date_of_birth",
    "date_of_birth": "date_of_birth",
    "last_name": "last_name",
    "first_name": "first_name",
    "phone_number": "phone_number",
    "position": "position",
}


@dataclass
class ParsedImportRow:
    """Строка Excel после нормализации."""

    row_number: int
    last_name: str
    first_name: str
    email: str
    phone_number: str
    position: str
    department: str
    date_of_birth: Optional[date] = None


@dataclass
class ImportRowResult:
    """Результат обработки одной строки."""

    row_number: int
    status: str
    last_name: str = ""
    first_name: str = ""
    email: str = ""
    phone_number: str = ""
    position: str = ""
    department: str = ""
    login: Optional[str] = None
    password: Optional[str] = None
    user_id: Optional[int] = None
    email_sent: bool = False
    message: str = ""


@dataclass
class BulkImportResult:
    """Итог массового импорта."""

    total_rows: int = 0
    created_count: int = 0
    skipped_count: int = 0
    error_count: int = 0
    emails_sent_count: int = 0
    rows: list[ImportRowResult] = field(default_factory=list)


def build_import_template_bytes() -> bytes:
    """Формирует пустой Excel-шаблон для загрузки сотрудников."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Сотрудники"
    worksheet.append(TEMPLATE_HEADERS)
    instructions = workbook.create_sheet("Инструкция")
    instructions.append(["Поле", "Обязательно", "Описание"])
    for item in [
        ("Фамилия", "Да", "Фамилия сотрудника"),
        ("Имя", "Да", "Имя сотрудника"),
        ("Email", "Да", "На этот адрес отправятся данные для входа"),
        ("Телефон", "Да", "Контактный номер"),
        ("Должность", "Да", "Должность сотрудника"),
        ("Ресторан", "Да", "Название ресторана / точки"),
        ("Дата рождения", "Нет", "Формат ДД.ММ.ГГГГ или ГГГГ-ММ-ДД"),
    ]:
        instructions.append(list(item))
    instructions.append([])
    instructions.append(["Пример строки (на листе «Сотрудники»):"])
    instructions.append(["Иванов", "Иван", "ivanov@restaurant.ru", "+79001234567", "Официант", "Москва-Стамбул", "15.03.1990"])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\*+$", "", text).strip()
    return text


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


def _parse_birth_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _map_headers(header_row: tuple[Any, ...]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for index, header in enumerate(header_row, start=1):
        normalized = _normalize_header(header)
        if not normalized:
            continue
        field_name = HEADER_ALIASES.get(normalized)
        if field_name:
            mapping[index] = field_name
    return mapping


def _find_data_sheet(workbook) -> Worksheet:
    if "Сотрудники" in workbook.sheetnames:
        return workbook["Сотрудники"]
    return workbook.active


def parse_import_workbook(file_bytes: bytes) -> tuple[list[ParsedImportRow], list[str]]:
    """Читает Excel и возвращает строки данных и ошибки формата."""
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    worksheet = _find_data_sheet(workbook)
    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], ["Файл пуст"]

    column_map = _map_headers(header_row)
    required_fields = {"last_name", "first_name", "email", "phone_number", "position", "department"}
    missing = required_fields - set(column_map.values())
    if missing:
        return [], [
            "Не найдены обязательные колонки. "
            f"Скачайте шаблон и заполните: {', '.join(TEMPLATE_HEADERS[:6])}",
        ]

    parsed_rows: list[ParsedImportRow] = []
    errors: list[str] = []
    for row_index, row in enumerate(rows_iter, start=2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        values: dict[str, Any] = {}
        for col_index, field_name in column_map.items():
            if col_index - 1 < len(row):
                values[field_name] = row[col_index - 1]
        last_name = _cell_text(values.get("last_name"))
        first_name = _cell_text(values.get("first_name"))
        if not last_name and not first_name:
            continue
        parsed_rows.append(
            ParsedImportRow(
                row_number=row_index,
                last_name=last_name,
                first_name=first_name,
                email=_cell_text(values.get("email")).lower(),
                phone_number=_cell_text(values.get("phone_number")),
                position=_cell_text(values.get("position")),
                department=_cell_text(values.get("department")),
                date_of_birth=_parse_birth_date(values.get("date_of_birth")),
            ),
        )
    if not parsed_rows:
        errors.append("В файле нет заполненных строк сотрудников")
    return parsed_rows, errors


def _validate_row(row: ParsedImportRow) -> Optional[str]:
    if not row.last_name:
        return "Не указана фамилия"
    if not row.first_name:
        return "Не указано имя"
    if not row.email:
        return "Не указан email"
    if not row.phone_number:
        return "Не указан телефон"
    if not row.position:
        return "Не указана должность"
    if not row.department:
        return "Не указан ресторан"
    if "@" not in row.email:
        return "Некорректный email"
    return None


async def _find_existing_user(db: AsyncSession, email: str, phone_number: str) -> Optional[models.User]:
    if email:
        result = await db.execute(
            select(models.User).where(
                models.User.email == email,
                models.User.status != "deleted",
            ),
        )
        user = result.scalar_one_or_none()
        if user:
            return user
    result = await db.execute(
        select(models.User).where(
            models.User.phone_number == phone_number,
            models.User.status != "deleted",
        ),
    )
    return result.scalar_one_or_none()


async def _create_imported_user(
    db: AsyncSession,
    row: ParsedImportRow,
    *,
    send_credentials: bool,
) -> ImportRowResult:
    validation_error = _validate_row(row)
    if validation_error:
        return ImportRowResult(
            row_number=row.row_number,
            status="error",
            last_name=row.last_name,
            first_name=row.first_name,
            email=row.email,
            phone_number=row.phone_number,
            position=row.position,
            department=row.department,
            message=validation_error,
        )

    existing = await _find_existing_user(db, row.email, row.phone_number)
    if existing:
        return ImportRowResult(
            row_number=row.row_number,
            status="skipped",
            last_name=row.last_name,
            first_name=row.first_name,
            email=row.email,
            phone_number=row.phone_number,
            position=row.position,
            department=row.department,
            message="Пользователь с таким email или телефоном уже существует",
        )

    db_user = models.User(
        telegram_id=None,
        first_name=row.first_name,
        last_name=row.last_name,
        position=row.position,
        department=row.department,
        phone_number=row.phone_number,
        date_of_birth=row.date_of_birth,
        email=row.email,
        status="approved",
        browser_auth_enabled=True,
        last_login_date=date.today(),
    )
    db.add(db_user)
    await db.commit()
    await db.refresh(db_user)

    base_login = generate_login_from_name(db_user.first_name, db_user.last_name, db_user.id)
    login = await _ensure_unique_login(db, base_login, db_user.id)
    plain_password = generate_random_password(12)
    db_user.login = login
    db_user.password_hash = get_password_hash(plain_password)
    db_user.password_plain = plain_password
    await db.commit()
    await db.refresh(db_user)

    email_sent = False
    if send_credentials and db_user.email:
        try:
            user_name = f"{db_user.first_name or ''} {db_user.last_name or ''}".strip()
            login_url = getattr(settings, "WEB_APP_LOGIN_URL", None) or None
            email_sent = await send_credentials_to_user(
                user_email=db_user.email,
                user_name=user_name,
                login=login,
                password=plain_password,
                login_url=login_url,
            )
        except Exception as error:
            logger.exception("Не удалось отправить email user_id=%s: %s", db_user.id, error)

    return ImportRowResult(
        row_number=row.row_number,
        status="created",
        last_name=row.last_name,
        first_name=row.first_name,
        email=row.email,
        phone_number=row.phone_number,
        position=row.position,
        department=row.department,
        login=login,
        password=plain_password,
        user_id=db_user.id,
        email_sent=email_sent,
        message="Создан" if email_sent else "Создан, email не отправлен (проверьте SMTP)",
    )


async def bulk_import_users(
    db: AsyncSession,
    rows: list[ParsedImportRow],
    *,
    send_credentials: bool = True,
) -> BulkImportResult:
    """Создаёт пользователей из распарсенных строк Excel."""
    result = BulkImportResult(total_rows=len(rows))
    for row in rows:
        row_result = await _create_imported_user(db, row, send_credentials=send_credentials)
        result.rows.append(row_result)
        if row_result.status == "created":
            result.created_count += 1
            if row_result.email_sent:
                result.emails_sent_count += 1
        elif row_result.status == "skipped":
            result.skipped_count += 1
        else:
            result.error_count += 1
    return result


def build_import_report_bytes(rows: list[ImportRowResult]) -> bytes:
    """Формирует Excel-отчёт с логинами и паролями для управляющего."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Доступы сотрудников"
    report_rows = [
        {
            "Строка": item.row_number,
            "Статус": item.status,
            "Фамилия": item.last_name,
            "Имя": item.first_name,
            "Email": item.email,
            "Телефон": item.phone_number,
            "Должность": item.position,
            "Ресторан": item.department,
            "Логин": item.login or "",
            "Пароль": item.password or "",
            "Email отправлен": "Да" if item.email_sent else "Нет",
            "Комментарий": item.message,
        }
        for item in rows
    ]
    if report_rows:
        headers = list(report_rows[0].keys())
        worksheet.append(headers)
        for row in report_rows:
            worksheet.append([row.get(header) for header in headers])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
